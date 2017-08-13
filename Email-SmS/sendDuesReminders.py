#coding=utf-8
#查找表格中特定人员，发送邮件

import openpyxl,smtplib
from email.mime.text import MIMEText

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

#查看表格最后列的内容
lactCol = sheet.max_column
latestMonth = sheet.cell(row=1,column=lactCol).value

unpifMembers = {}
for r in range(2,sheet.max_row + 1):
	payment = sheet.cell(row=r,column=lactCol).value
	if payment != 'paid':
		name = sheet.cell(row=r,column=1).value
		email = sheet.cell(row=r,column=2).value
		unpifMembers[name] = email
print unpifMembers

#设置发送邮件信息
smtpObj=smtplib.SMTP_SSL('smtp.163.com',465)
smtpObj.ehlo()
# smtpObj.starttls()
sender='XXXXo@163.com'
password='XXXX'
smtpObj.login(sender,password)


for name,email in unpifMembers.items():
	#设置邮件发送格式，包含邮件内容，主题，发信人，收信人
	count='Dear %s,\nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!'%(name,latestMonth)
	msg = MIMEText(count, _subtype='plain', _charset='utf-8')
	msg['From'] = sender
	msg['Subject']= '%s dues unpaid.'%latestMonth
	msg['To'] = email
	# body = 'Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!' % (
	# latestMonth, name, latestMonth)
	print('Sending email to %s...' % email)
	#发送邮件
	sendmailStatus = smtpObj.sendmail(sender,email,	msg.as_string())

	if sendmailStatus !={}:
		print 'There was a problem sending email to %s:%s'%(email,sendmailStatus)

smtpObj.quit()



