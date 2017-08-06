#coding=utf-8


import  openpyxl,sys
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font,NamedStyle

#获取命令行参数
if len(sys.argv) < 2:
	print 'usage: py multiplicationTable.py [NUM]'
	exit()

#程序名是0，故参数为1
num = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()

for i in range(1,num+1):
	sheet[get_column_letter(i+1)+'1'] = i
	sheet['A'+str(i+1)] = i

for i in range(1,num+1):
	for j in range(1,num+1):
		sheet[get_column_letter(j+1)+str(i+1)] = i*j


wb.save('MultiplicationTable-'+str(num)+'.xlsx')