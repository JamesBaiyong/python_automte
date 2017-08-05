#coding=utf-8
#更新电子表格中的数据

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

#定义字典，使用字典的内容作为更新值，比使用IF语句更方便优雅
PRICE_UPDATES={
	'Garlic':3.08,
	'Celery':1.20,
	'Lemon':1.26
}

#遍历每一行，查找需要更新的单元格更新数据
for rowNum in range(2,sheet.max_row):
	# 把第一列单元格的值保存在变量中
	produceName = sheet.cell(row=rowNum,column=1).value
	# 如果变量的值是字典中的键，就更新值
	if produceName in PRICE_UPDATES:
		sheet.cell(row=rowNum,column=2).value = PRICE_UPDATES[produceName]
wb.save('updatedProduceSales.xlsx')
